import json
import os

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "skill__documents__create_pdf",
            "description": "Create a PDF file with a title and plain-text body.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Absolute output path"},
                    "title": {"type": "string", "description": "Document title (bold, centered)"},
                    "body": {"type": "string", "description": "Body text; newlines preserved"},
                },
                "required": ["path", "title", "body"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "skill__documents__create_excel",
            "description": "Create an .xlsx file. Pass headers and rows as JSON strings.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Absolute output path"},
                    "title": {"type": "string", "description": "Sheet title (<=31 chars)"},
                    "headers_json": {
                        "type": "string",
                        "description": 'JSON array of column headers, e.g. ["Name","Age"]',
                    },
                    "rows_json": {
                        "type": "string",
                        "description": 'JSON array of row arrays, e.g. [["Ada",36]]',
                    },
                },
                "required": ["path", "title", "headers_json", "rows_json"],
            },
        },
    },
]


def execute(name, args, ctx):
    if name == "skill__documents__create_pdf":
        return _create_pdf(
            os.path.expanduser(args.get("path", "")),
            args.get("title", ""),
            args.get("body", ""),
        )
    if name == "skill__documents__create_excel":
        return _create_excel(
            os.path.expanduser(args.get("path", "")),
            args.get("title", ""),
            args.get("headers_json", "[]"),
            args.get("rows_json", "[]"),
        )
    return f"Error: unknown tool {name}"


def _ensure_parent(path):
    directory = os.path.dirname(path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)


def _create_pdf(path, title, body):
    if not path:
        return "Error: path is required"
    from fpdf import FPDF

    _ensure_parent(path)
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, title, ln=True, align="C")
    pdf.ln(6)
    pdf.set_font("Helvetica", "", 11)
    for line in body.split("\n"):
        pdf.multi_cell(0, 6, line)
    pdf.output(path)
    return f"PDF created: {path} ({len(body)} chars body)"


def _create_excel(path, title, headers_json, rows_json):
    if not path:
        return "Error: path is required"
    from openpyxl import Workbook
    from openpyxl.styles import Font

    try:
        headers = json.loads(headers_json) if headers_json else []
        rows = json.loads(rows_json) if rows_json else []
    except json.JSONDecodeError as e:
        return f"Error: invalid JSON: {e}"
    if not isinstance(headers, list) or not isinstance(rows, list):
        return "Error: headers_json and rows_json must be JSON arrays"

    _ensure_parent(path)
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet1")[:31]
    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return f"Excel created: {path} ({len(rows)} rows)"
