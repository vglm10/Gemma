import subprocess
import os
import json
import glob as glob_module
from datetime import datetime

MAX_RESULT_CHARS = 10000

TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "run_command",
            "description": "Execute a shell command on the user's macOS computer and return stdout/stderr. Use for system tasks, checking status, installing software, running scripts, etc.",
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {
                        "type": "string",
                        "description": "The shell command to execute",
                    }
                },
                "required": ["command"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "Read the contents of a file. Returns the text content.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Absolute path to the file to read",
                    }
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": "Write content to a file. Creates the file if it doesn't exist, overwrites if it does.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Absolute path to the file to write",
                    },
                    "content": {
                        "type": "string",
                        "description": "The content to write to the file",
                    },
                },
                "required": ["path", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_directory",
            "description": "List files and folders in a directory with sizes and modification dates.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Absolute path to the directory to list",
                    }
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": "Search for files matching a glob pattern. Returns matching file paths.",
            "parameters": {
                "type": "object",
                "properties": {
                    "pattern": {
                        "type": "string",
                        "description": "Glob pattern to search for, e.g. '~/Documents/**/*.pdf'",
                    }
                },
                "required": ["pattern"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_document",
            "description": "Create a PDF or Excel document. For PDF, provide title and text content. For Excel, provide headers and rows of data.",
            "parameters": {
                "type": "object",
                "properties": {
                    "format": {
                        "type": "string",
                        "enum": ["pdf", "excel"],
                        "description": "Document format: 'pdf' or 'excel'",
                    },
                    "path": {
                        "type": "string",
                        "description": "Absolute path where the document will be saved",
                    },
                    "title": {
                        "type": "string",
                        "description": "Document title",
                    },
                    "content": {
                        "type": "string",
                        "description": "For PDF: the text content. For Excel: JSON string with 'headers' (list) and 'rows' (list of lists).",
                    },
                },
                "required": ["format", "path", "title", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "manage_schedule",
            "description": "Manage scheduled tasks. Actions: 'create' (set up a recurring task), 'list' (show all tasks), 'delete' (remove a task by id).",
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["create", "list", "delete"],
                        "description": "The action to perform",
                    },
                    "name": {
                        "type": "string",
                        "description": "Task name (for 'create')",
                    },
                    "prompt": {
                        "type": "string",
                        "description": "The prompt/instruction to run on schedule (for 'create')",
                    },
                    "interval_minutes": {
                        "type": "number",
                        "description": "How often to run in minutes (for 'create')",
                    },
                    "task_id": {
                        "type": "string",
                        "description": "Task ID to delete (for 'delete')",
                    },
                },
                "required": ["action"],
            },
        },
    },
]

# Scheduler reference — set by main.py at startup
_scheduler = None


def set_scheduler(scheduler):
    global _scheduler
    _scheduler = scheduler


def _truncate(text):
    if len(text) > MAX_RESULT_CHARS:
        return text[:MAX_RESULT_CHARS] + f"\n\n[Truncated — {len(text)} total chars]"
    return text


def execute_tool(name, args):
    """Execute a tool by name with given args. Returns result string."""
    try:
        if name == "run_command":
            return _run_command(args.get("command", ""))
        elif name == "read_file":
            return _read_file(args.get("path", ""))
        elif name == "write_file":
            return _write_file(args.get("path", ""), args.get("content", ""))
        elif name == "list_directory":
            return _list_directory(args.get("path", ""))
        elif name == "search_files":
            return _search_files(args.get("pattern", ""))
        elif name == "create_document":
            return _create_document(
                args.get("format", "pdf"),
                args.get("path", ""),
                args.get("title", ""),
                args.get("content", ""),
            )
        elif name == "manage_schedule":
            return _manage_schedule(args)
        else:
            return f"Unknown tool: {name}"
    except Exception as e:
        return f"Error: {e}"


def _run_command(command):
    if not command.strip():
        return "Error: empty command"
    try:
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            timeout=30,
            env={**os.environ, "PATH": f"/opt/homebrew/bin:/usr/local/bin:{os.environ.get('PATH', '')}"},
        )
        output = result.stdout
        if result.stderr:
            output += ("\n" if output else "") + result.stderr
        if result.returncode != 0:
            output += f"\n[Exit code: {result.returncode}]"
        return _truncate(output.strip() or "(no output)")
    except subprocess.TimeoutExpired:
        return "Error: Command timed out after 30 seconds"


def _read_file(path):
    path = os.path.expanduser(path)
    if not os.path.exists(path):
        return f"Error: File not found: {path}"
    if os.path.isdir(path):
        return f"Error: {path} is a directory, not a file. Use list_directory instead."
    size = os.path.getsize(path)
    if size > 100_000:
        return f"Error: File too large ({size} bytes, max 100KB)"
    with open(path, "r", errors="replace") as f:
        return _truncate(f.read())


def _write_file(path, content):
    path = os.path.expanduser(path)
    directory = os.path.dirname(path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
    with open(path, "w") as f:
        f.write(content)
    return f"File written: {path} ({len(content)} chars)"


def _list_directory(path):
    path = os.path.expanduser(path)
    if not os.path.exists(path):
        return f"Error: Directory not found: {path}"
    if not os.path.isdir(path):
        return f"Error: {path} is not a directory"
    entries = []
    try:
        items = sorted(os.listdir(path))
    except PermissionError:
        return f"Error: Permission denied: {path}"
    for item in items[:200]:
        full = os.path.join(path, item)
        try:
            stat = os.stat(full)
            is_dir = os.path.isdir(full)
            size = stat.st_size if not is_dir else "-"
            mtime = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")
            prefix = "d" if is_dir else "f"
            size_str = _human_size(stat.st_size) if not is_dir else "   -"
            entries.append(f"  {prefix}  {size_str:>8}  {mtime}  {item}")
        except (PermissionError, OSError):
            entries.append(f"  ?  {'?':>8}  {'?':>16}  {item}")
    header = f"Directory: {path} ({len(items)} items)\n"
    if len(items) > 200:
        header += f"(showing first 200 of {len(items)})\n"
    return header + "\n".join(entries)


def _human_size(size):
    for unit in ("B", "KB", "MB", "GB"):
        if abs(size) < 1024:
            return f"{size:.0f}{unit}" if unit == "B" else f"{size:.1f}{unit}"
        size /= 1024
    return f"{size:.1f}TB"


def _search_files(pattern):
    pattern = os.path.expanduser(pattern)
    matches = glob_module.glob(pattern, recursive=True)
    if not matches:
        return "No files found matching pattern."
    matches = matches[:100]
    result = f"Found {len(matches)} file(s):\n"
    result += "\n".join(f"  {m}" for m in matches)
    return result


def _create_document(fmt, path, title, content):
    path = os.path.expanduser(path)
    directory = os.path.dirname(path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)

    if fmt == "pdf":
        return _create_pdf(path, title, content)
    elif fmt == "excel":
        return _create_excel(path, title, content)
    else:
        return f"Error: Unknown format '{fmt}'. Use 'pdf' or 'excel'."


def _create_pdf(path, title, content):
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, title, ln=True, align="C")
    pdf.ln(6)
    pdf.set_font("Helvetica", "", 11)
    for line in content.split("\n"):
        pdf.multi_cell(0, 6, line)
    pdf.output(path)
    return f"PDF created: {path}"


def _create_excel(path, title, content):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        return "Error: Excel content must be a JSON string with 'headers' and 'rows'"

    headers = data.get("headers", [])
    rows = data.get("rows", [])

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)

    wb.save(path)
    return f"Excel file created: {path} ({len(rows)} rows)"


def _manage_schedule(args):
    action = args.get("action", "")
    if _scheduler is None:
        return "Error: Scheduler not initialized"

    if action == "create":
        name = args.get("name", "Untitled")
        prompt = args.get("prompt", "")
        interval = args.get("interval_minutes", 60)
        if not prompt:
            return "Error: 'prompt' is required to create a scheduled task"
        task = _scheduler.add_task(name, prompt, int(interval))
        return f"Scheduled task created: '{name}' (every {interval} min, id: {task['id']})"

    elif action == "list":
        tasks = _scheduler.get_tasks()
        if not tasks:
            return "No scheduled tasks."
        lines = ["Scheduled tasks:"]
        for t in tasks:
            status = "enabled" if t["enabled"] else "disabled"
            last = datetime.fromtimestamp(t["last_run"]).strftime("%H:%M") if t["last_run"] else "never"
            lines.append(f"  - {t['name']} (every {t['interval_minutes']}min, {status}, last: {last}, id: {t['id']})")
        return "\n".join(lines)

    elif action == "delete":
        task_id = args.get("task_id", "")
        if _scheduler.remove_task(task_id):
            return f"Task {task_id} deleted."
        return f"Error: Task {task_id} not found."

    return f"Error: Unknown action '{action}'"
